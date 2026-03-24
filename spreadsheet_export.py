from __future__ import annotations

import os
import site
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

workspace_packages = (Path(__file__).resolve().parent / ".python_packages").resolve()
if workspace_packages.exists():
    sys.path.append(str(workspace_packages))

user_site = site.getusersitepackages()
if user_site and user_site not in sys.path:
    sys.path.append(user_site)

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    Workbook = None
    BarChart = Reference = CellIsRule = Alignment = Border = Font = PatternFill = Side = None
    Table = TableStyleInfo = None
    get_column_letter = None


DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"
BAND_FORMAT = "0.0"
WHOLE_NUMBER_FORMAT = "0"

if Workbook is not None:
    HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    TITLE_FILL = PatternFill("solid", fgColor="16324F")
    TITLE_FONT = Font(bold=True, color="FFFFFF", size=15)
    LABEL_FILL = PatternFill("solid", fgColor="F2F4F7")
    KPI_FILL = PatternFill("solid", fgColor="D9F3F4")
    CAUTION_FILL = PatternFill("solid", fgColor="FFF0D9")
    SUCCESS_FILL = PatternFill("solid", fgColor="E6F7EA")
    IMPORTED_FONT = Font(color="1E7D32")
    FORMULA_FONT = Font(color="000000")
    STATIC_FONT = Font(color="666666")
    LINK_FONT = Font(color="1E7D32", underline="single")
    SUBTLE_BORDER = Border(bottom=Side(style="thin", color="D0D7DE"))
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT_WRAP = Alignment(vertical="top", wrap_text=True)
else:
    HEADER_FILL = HEADER_FONT = TITLE_FILL = TITLE_FONT = None
    LABEL_FILL = KPI_FILL = CAUTION_FILL = SUCCESS_FILL = None
    IMPORTED_FONT = FORMULA_FONT = STATIC_FONT = LINK_FONT = None
    SUBTLE_BORDER = CENTER = LEFT_WRAP = None


def export_database_to_workbook(database_path: str | Path, workbook_path: str | Path) -> Path | None:
    if Workbook is None:
        return None

    db_path = Path(database_path).resolve()
    output_path = Path(workbook_path).resolve()
    temp_dir = (db_path.parent / "tmp" / "spreadsheets").resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_dir.mkdir(parents=True, exist_ok=True)

    payload = _read_database(db_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    export_time = datetime.now()
    _build_dashboard_sheet(workbook, export_time)
    _build_candidates_sheet(workbook, payload["candidates"], db_path)
    _build_submissions_sheet(workbook, payload["submissions"], db_path)
    _build_legacy_users_sheet(workbook, payload["legacy_users"], db_path)
    _build_legacy_attempts_sheet(workbook, payload["legacy_attempts"], db_path)
    _build_settings_sheet(workbook, payload["settings"], db_path)
    _build_sources_sheet(workbook, db_path, output_path, export_time)

    temp_path = temp_dir / f"{output_path.stem}.tmp.xlsx"
    workbook.save(temp_path)
    os.replace(temp_path, output_path)
    return output_path


def _read_database(db_path: Path) -> dict[str, list[dict[str, Any]]]:
    connection = sqlite3.connect(db_path)
    connection.row_factory = sqlite3.Row
    try:
        tables = {row["name"] for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'")}
        candidates = []
        submissions = []
        legacy_users = []
        legacy_attempts = []
        settings = []

        if "candidates" in tables:
            candidates = [
                dict(row)
                for row in connection.execute(
                    "SELECT id, candidate_code, full_name, email, age, country, target_band, created_at FROM candidates ORDER BY id"
                )
            ]
        if "submissions" in tables and "candidates" in tables:
            submissions = [
                dict(row)
                for row in connection.execute(
                    """
                    SELECT
                        submissions.id,
                        submissions.candidate_id,
                        candidates.candidate_code,
                        candidates.full_name,
                        candidates.target_band,
                        submissions.status,
                        submissions.current_step,
                        submissions.listening_correct,
                        submissions.listening_band,
                        submissions.reading_correct,
                        submissions.reading_band,
                        submissions.writing_band,
                        submissions.speaking_band,
                        submissions.overall_band,
                        submissions.position_label,
                        submissions.candidate_notes,
                        submissions.examiner_feedback,
                        submissions.speaking_part_1_audio,
                        submissions.speaking_part_2_audio,
                        submissions.speaking_part_3_audio,
                        submissions.created_at,
                        submissions.submitted_at,
                        submissions.scored_at,
                        submissions.updated_at
                    FROM submissions
                    JOIN candidates ON candidates.id = submissions.candidate_id
                    ORDER BY submissions.id
                    """
                )
            ]
        if "users" in tables:
            legacy_users = [dict(row) for row in connection.execute("SELECT * FROM users ORDER BY id")]
        if "attempts" in tables and "users" in tables:
            legacy_attempts = [
                dict(row)
                for row in connection.execute(
                    """
                    SELECT
                        attempts.id,
                        attempts.user_id,
                        users.full_name,
                        users.email,
                        users.target_band,
                        attempts.listening_raw,
                        attempts.reading_raw,
                        attempts.writing_band,
                        attempts.speaking_band,
                        attempts.listening_band,
                        attempts.reading_band,
                        attempts.overall_band,
                        attempts.position_label,
                        attempts.notes,
                        attempts.created_at
                    FROM attempts
                    JOIN users ON users.id = attempts.user_id
                    ORDER BY attempts.id
                    """
                )
            ]
        if "app_settings" in tables:
            settings = [dict(row) for row in connection.execute("SELECT key, value FROM app_settings ORDER BY key")]
    finally:
        connection.close()

    return {
        "candidates": candidates,
        "submissions": submissions,
        "legacy_users": legacy_users,
        "legacy_attempts": legacy_attempts,
        "settings": settings,
    }


def _build_dashboard_sheet(workbook, export_time: datetime) -> None:
    sheet = workbook.create_sheet("Dashboard")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "BandForge Data Workbook"
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].alignment = CENTER
    sheet["A2"] = "A single workbook for current candidate submissions, app settings, and any legacy IELTS records."
    sheet["A2"].font = STATIC_FONT

    _write_row(
        sheet,
        4,
        [
            ("Metric", HEADER_FILL, HEADER_FONT, CENTER),
            ("Value", HEADER_FILL, HEADER_FONT, CENTER),
        ],
    )
    metrics = [
        ("Exported at", export_time),
        ("Public access", '=IF(Settings!B2="on","ON","OFF")'),
        ("Current candidates", '=MAX(COUNTA(Candidates!A:A)-1,0)'),
        ("Current submissions", '=MAX(COUNTA(Submissions!A:A)-1,0)'),
        ("Published results", '=COUNTIF(Submissions!E:E,"published")'),
        ("Average overall band", '=IF(COUNT(Submissions!O:O)=0,"",AVERAGE(Submissions!O:O))'),
        ("Best overall band", '=IF(COUNT(Submissions!O:O)=0,"",MAX(Submissions!O:O))'),
        ("Candidates meeting target", '=COUNTIF(Candidates!L:L,"<=0")'),
        ("Legacy users", "=MAX(COUNTA('Legacy Users'!A:A)-1,0)"),
        ("Legacy attempts", "=MAX(COUNTA('Legacy Attempts'!A:A)-1,0)"),
        ("Legacy average overall", '=IF(COUNT(\'Legacy Attempts\'!L:L)=0,"",AVERAGE(\'Legacy Attempts\'!L:L))'),
    ]
    for row_index, (label, value) in enumerate(metrics, start=5):
        label_cell = sheet.cell(row=row_index, column=1, value=label)
        value_cell = sheet.cell(row=row_index, column=2, value=value)
        label_cell.fill = LABEL_FILL
        label_cell.border = SUBTLE_BORDER
        value_cell.fill = KPI_FILL
        value_cell.border = SUBTLE_BORDER
        value_cell.font = FORMULA_FONT
        if isinstance(value, datetime):
            value_cell.number_format = DATETIME_FORMAT
        elif row_index in {10, 11, 15}:
            value_cell.number_format = BAND_FORMAT
        elif row_index == 5:
            value_cell.number_format = DATETIME_FORMAT

    _write_row(
        sheet,
        4,
        [
            ("Section", HEADER_FILL, HEADER_FONT, CENTER),
            ("Average Band", HEADER_FILL, HEADER_FONT, CENTER),
        ],
        start_column=4,
    )
    sections = [
        ("Listening", '=IF(COUNT(Submissions!I:I)=0,"",AVERAGE(Submissions!I:I))'),
        ("Reading", '=IF(COUNT(Submissions!K:K)=0,"",AVERAGE(Submissions!K:K))'),
        ("Writing", '=IF(COUNT(Submissions!L:L)=0,"",AVERAGE(Submissions!L:L))'),
        ("Speaking", '=IF(COUNT(Submissions!M:M)=0,"",AVERAGE(Submissions!M:M))'),
        ("Average gap to target", '=IF(COUNT(Submissions!Q:Q)=0,"",AVERAGE(Submissions!Q:Q))'),
    ]
    for row_index, (label, formula) in enumerate(sections, start=5):
        label_cell = sheet.cell(row=row_index, column=4, value=label)
        value_cell = sheet.cell(row=row_index, column=5, value=formula)
        label_cell.fill = LABEL_FILL
        label_cell.border = SUBTLE_BORDER
        value_cell.fill = KPI_FILL
        value_cell.border = SUBTLE_BORDER
        value_cell.font = FORMULA_FONT
        value_cell.number_format = BAND_FORMAT

    _write_row(
        sheet,
        11,
        [
            ("Position", HEADER_FILL, HEADER_FONT, CENTER),
            ("Count", HEADER_FILL, HEADER_FONT, CENTER),
        ],
        start_column=4,
    )
    labels = [
        "Exam Ready",
        "Strong Progress",
        "Developing Competence",
        "Emerging Control",
        "Foundation Stage",
    ]
    for row_index, label in enumerate(labels, start=12):
        label_cell = sheet.cell(row=row_index, column=4, value=label)
        count_cell = sheet.cell(row=row_index, column=5, value=f'=COUNTIF(Submissions!P:P,D{row_index})')
        label_cell.fill = LABEL_FILL
        label_cell.border = SUBTLE_BORDER
        count_cell.fill = KPI_FILL
        count_cell.border = SUBTLE_BORDER
        count_cell.font = FORMULA_FONT
        count_cell.number_format = WHOLE_NUMBER_FORMAT

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Published Position Distribution"
    chart.y_axis.title = "Position"
    chart.x_axis.title = "Submissions"
    chart.height = 6
    chart.width = 10
    data = Reference(sheet, min_col=5, min_row=11, max_row=16)
    categories = Reference(sheet, min_col=4, min_row=12, max_row=16)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "G4")

    sheet["D18"] = "Workbook notes"
    sheet["E18"] = "Green text shows imported values, black text shows formulas, and gap columns use color to flag whether targets are met."
    sheet["D18"].fill = LABEL_FILL
    sheet["D18"].border = SUBTLE_BORDER
    sheet["E18"].fill = CAUTION_FILL
    sheet["E18"].border = SUBTLE_BORDER
    sheet["E18"].alignment = LEFT_WRAP

    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["D"].width = 24
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["G"].width = 18


def _build_candidates_sheet(workbook, rows: list[dict[str, Any]], db_path: Path) -> None:
    sheet = workbook.create_sheet("Candidates")
    headers = [
        "Candidate ID",
        "Candidate Code",
        "Full Name",
        "Email",
        "Age",
        "Country",
        "Target Band",
        "Created At",
        "Submission Count",
        "Published Count",
        "Best Overall Band",
        "Gap to Best Result",
        "Data Source",
    ]
    _write_headers(sheet, headers)
    source_url = db_path.as_uri()

    for row_index, row in enumerate(rows, start=2):
        values = [
            row["id"],
            row["candidate_code"],
            row["full_name"],
            row["email"],
            row["age"],
            row["country"],
            row["target_band"],
            _parse_datetime(row["created_at"]),
            f'=COUNTIF(Submissions!$B:$B,A{row_index})',
            f'=COUNTIFS(Submissions!$B:$B,A{row_index},Submissions!$E:$E,"published")',
            f'=IF(J{row_index}=0,"",MAXIFS(Submissions!$O:$O,Submissions!$B:$B,A{row_index},Submissions!$E:$E,"published"))',
            f'=IF(K{row_index}="","",G{row_index}-K{row_index})',
            source_url,
        ]
        _append_row(sheet, values)
        _style_imported_cells(sheet, row_index, [1, 2, 3, 4, 5, 6, 7, 8])
        _style_formula_cells(sheet, row_index, [9, 10, 11, 12])
        _style_link_cell(sheet.cell(row=row_index, column=13), source_url)

    _finalize_detail_sheet(sheet, len(headers), "CandidatesTable")
    widths = [12, 18, 24, 30, 10, 18, 12, 20, 16, 16, 16, 16, 34]
    _set_column_widths(sheet, widths)
    for row in range(2, sheet.max_row + 1):
        for column in [7, 11, 12]:
            sheet.cell(row=row, column=column).number_format = BAND_FORMAT
        sheet.cell(row=row, column=8).number_format = DATETIME_FORMAT
    _add_gap_highlighting(sheet, f"L2:L{sheet.max_row}")


def _build_submissions_sheet(workbook, rows: list[dict[str, Any]], db_path: Path) -> None:
    sheet = workbook.create_sheet("Submissions")
    headers = [
        "Submission ID",
        "Candidate ID",
        "Candidate Code",
        "Candidate Name",
        "Status",
        "Current Step",
        "Target Band",
        "Listening Correct",
        "Listening Band",
        "Reading Correct",
        "Reading Band",
        "Writing Band",
        "Speaking Band",
        "Section Average",
        "Overall Band",
        "Position Label",
        "Gap to Target",
        "Target Status",
        "Candidate Notes",
        "Examiner Feedback",
        "Audio 1",
        "Audio 2",
        "Audio 3",
        "Created At",
        "Submitted At",
        "Scored At",
        "Updated At",
        "Data Source",
    ]
    _write_headers(sheet, headers)
    source_url = db_path.as_uri()

    for row_index, row in enumerate(rows, start=2):
        values = [
            row["id"],
            row["candidate_id"],
            row["candidate_code"],
            row["full_name"],
            row["status"],
            row["current_step"],
            row["target_band"],
            row["listening_correct"],
            row["listening_band"],
            row["reading_correct"],
            row["reading_band"],
            row["writing_band"],
            row["speaking_band"],
            f'=IF(COUNT(I{row_index},K{row_index},L{row_index},M{row_index})<4,"",(I{row_index}+K{row_index}+L{row_index}+M{row_index})/4)',
            row["overall_band"],
            row["position_label"],
            f'=IF(OR(G{row_index}="",O{row_index}=""),"",G{row_index}-O{row_index})',
            f'=IF(Q{row_index}="","",IF(Q{row_index}<=0,"At or Above Target","Below Target"))',
            row["candidate_notes"],
            row["examiner_feedback"],
            row["speaking_part_1_audio"],
            row["speaking_part_2_audio"],
            row["speaking_part_3_audio"],
            _parse_datetime(row["created_at"]),
            _parse_datetime(row["submitted_at"]),
            _parse_datetime(row["scored_at"]),
            _parse_datetime(row["updated_at"]),
            source_url,
        ]
        _append_row(sheet, values)
        _style_imported_cells(sheet, row_index, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 19, 20, 21, 22, 23, 24, 25, 26, 27])
        _style_formula_cells(sheet, row_index, [14, 17, 18])
        _style_link_cell(sheet.cell(row=row_index, column=28), source_url)

    _finalize_detail_sheet(sheet, len(headers), "SubmissionsTable")
    widths = [12, 12, 18, 24, 14, 14, 12, 16, 13, 16, 13, 13, 13, 14, 12, 20, 12, 18, 26, 26, 18, 18, 18, 20, 20, 20, 20, 34]
    _set_column_widths(sheet, widths)
    for row in range(2, sheet.max_row + 1):
        for column in [7, 9, 11, 12, 13, 14, 15, 17]:
            sheet.cell(row=row, column=column).number_format = BAND_FORMAT
        for column in [24, 25, 26, 27]:
            sheet.cell(row=row, column=column).number_format = DATETIME_FORMAT
        for column in [19, 20]:
            sheet.cell(row=row, column=column).alignment = LEFT_WRAP
    _add_gap_highlighting(sheet, f"Q2:Q{sheet.max_row}")


def _build_legacy_users_sheet(workbook, rows: list[dict[str, Any]], db_path: Path) -> None:
    sheet = workbook.create_sheet("Legacy Users")
    headers = [
        "User ID",
        "Full Name",
        "Email",
        "Age",
        "Country",
        "Target Band",
        "Created At",
        "Attempt Count",
        "Best Overall Band",
        "Gap to Best Result",
        "Data Source",
    ]
    _write_headers(sheet, headers)
    source_url = db_path.as_uri()

    for row_index, row in enumerate(rows, start=2):
        values = [
            row["id"],
            row["full_name"],
            row["email"],
            row["age"],
            row["country"],
            row["target_band"],
            _parse_datetime(row["created_at"]),
            f'=COUNTIF(\'Legacy Attempts\'!$B:$B,A{row_index})',
            f'=IF(H{row_index}=0,"",MAXIFS(\'Legacy Attempts\'!$L:$L,\'Legacy Attempts\'!$B:$B,A{row_index}))',
            f'=IF(I{row_index}="","",F{row_index}-I{row_index})',
            source_url,
        ]
        _append_row(sheet, values)
        _style_imported_cells(sheet, row_index, [1, 2, 3, 4, 5, 6, 7])
        _style_formula_cells(sheet, row_index, [8, 9, 10])
        _style_link_cell(sheet.cell(row=row_index, column=11), source_url)

    _finalize_detail_sheet(sheet, len(headers), "LegacyUsersTable")
    widths = [10, 24, 30, 10, 18, 12, 20, 14, 16, 16, 34]
    _set_column_widths(sheet, widths)
    for row in range(2, sheet.max_row + 1):
        for column in [6, 9, 10]:
            sheet.cell(row=row, column=column).number_format = BAND_FORMAT
        sheet.cell(row=row, column=7).number_format = DATETIME_FORMAT
    _add_gap_highlighting(sheet, f"J2:J{sheet.max_row}")


def _build_legacy_attempts_sheet(workbook, rows: list[dict[str, Any]], db_path: Path) -> None:
    sheet = workbook.create_sheet("Legacy Attempts")
    headers = [
        "Attempt ID",
        "User ID",
        "Full Name",
        "Email",
        "Target Band",
        "Listening Raw",
        "Reading Raw",
        "Listening Band",
        "Reading Band",
        "Writing Band",
        "Speaking Band",
        "Overall Band",
        "Position Label",
        "Gap to Target",
        "Target Status",
        "Notes",
        "Created At",
        "Data Source",
    ]
    _write_headers(sheet, headers)
    source_url = db_path.as_uri()

    for row_index, row in enumerate(rows, start=2):
        values = [
            row["id"],
            row["user_id"],
            row["full_name"],
            row["email"],
            row["target_band"],
            row["listening_raw"],
            row["reading_raw"],
            row["listening_band"],
            row["reading_band"],
            row["writing_band"],
            row["speaking_band"],
            row["overall_band"],
            row["position_label"],
            f'=IF(OR(E{row_index}="",L{row_index}=""),"",E{row_index}-L{row_index})',
            f'=IF(N{row_index}="","",IF(N{row_index}<=0,"At or Above Target","Below Target"))',
            row["notes"],
            _parse_datetime(row["created_at"]),
            source_url,
        ]
        _append_row(sheet, values)
        _style_imported_cells(sheet, row_index, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 16, 17])
        _style_formula_cells(sheet, row_index, [14, 15])
        _style_link_cell(sheet.cell(row=row_index, column=18), source_url)

    _finalize_detail_sheet(sheet, len(headers), "LegacyAttemptsTable")
    widths = [10, 10, 24, 30, 12, 12, 12, 13, 13, 13, 13, 12, 18, 12, 18, 24, 20, 34]
    _set_column_widths(sheet, widths)
    for row in range(2, sheet.max_row + 1):
        for column in [5, 8, 9, 10, 11, 12, 14]:
            sheet.cell(row=row, column=column).number_format = BAND_FORMAT
        sheet.cell(row=row, column=17).number_format = DATETIME_FORMAT
        sheet.cell(row=row, column=16).alignment = LEFT_WRAP
    _add_gap_highlighting(sheet, f"N2:N{sheet.max_row}")


def _build_settings_sheet(workbook, rows: list[dict[str, Any]], db_path: Path) -> None:
    sheet = workbook.create_sheet("Settings")
    headers = ["Key", "Value", "Description", "Data Source"]
    _write_headers(sheet, headers)
    source_url = db_path.as_uri()
    descriptions = {
        "public_access": "Controls whether the public exam flow is reachable from the website.",
    }

    for row_index, row in enumerate(rows, start=2):
        values = [
            row["key"],
            row["value"],
            descriptions.get(row["key"], "Application setting synced from SQLite."),
            source_url,
        ]
        _append_row(sheet, values)
        _style_imported_cells(sheet, row_index, [1, 2, 3])
        _style_link_cell(sheet.cell(row=row_index, column=4), source_url)

    _finalize_detail_sheet(sheet, len(headers), "SettingsTable")
    _set_column_widths(sheet, [18, 14, 58, 34])


def _build_sources_sheet(workbook, db_path: Path, workbook_path: Path, export_time: datetime) -> None:
    sheet = workbook.create_sheet("Sources")
    headers = ["Item", "Location", "Notes"]
    _write_headers(sheet, headers)
    rows = [
        ("Database", db_path.as_uri(), "Primary source for application data and spreadsheet exports."),
        ("Application logic", (db_path.parent / "app.py").resolve().as_uri(), "Contains schema creation, scoring rules, and export triggers."),
        ("Workbook output", workbook_path.as_uri(), f"Latest workbook generated on {export_time.strftime('%Y-%m-%d %H:%M:%S')}."),
    ]
    for row_index, row in enumerate(rows, start=2):
        _append_row(sheet, list(row))
        _style_imported_cells(sheet, row_index, [1, 3])
        _style_link_cell(sheet.cell(row=row_index, column=2), row[1])

    _finalize_detail_sheet(sheet, len(headers), "SourcesTable")
    _set_column_widths(sheet, [24, 44, 56])


def _write_headers(sheet, headers: list[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = SUBTLE_BORDER
        cell.alignment = CENTER if len(header) <= 16 else LEFT_WRAP


def _write_row(sheet, row_index: int, values: list[tuple[Any, Any, Any, Any]], start_column: int = 1) -> None:
    for offset, (value, fill, font, alignment) in enumerate(values):
        cell = sheet.cell(row=row_index, column=start_column + offset, value=value)
        cell.fill = fill
        cell.font = font
        cell.border = SUBTLE_BORDER
        cell.alignment = alignment


def _append_row(sheet, values: list[Any]) -> None:
    sheet.append(values)


def _style_imported_cells(sheet, row_index: int, columns: list[int]) -> None:
    for column in columns:
        cell = sheet.cell(row=row_index, column=column)
        cell.font = IMPORTED_FONT
        cell.border = SUBTLE_BORDER


def _style_formula_cells(sheet, row_index: int, columns: list[int]) -> None:
    for column in columns:
        cell = sheet.cell(row=row_index, column=column)
        cell.font = FORMULA_FONT
        cell.border = SUBTLE_BORDER


def _style_link_cell(cell, url: str) -> None:
    cell.value = url
    cell.hyperlink = url
    cell.font = LINK_FONT
    cell.border = SUBTLE_BORDER
    cell.alignment = LEFT_WRAP


def _finalize_detail_sheet(sheet, column_count: int, table_name: str) -> None:
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 2:
        ref = f"A1:{get_column_letter(column_count)}{sheet.max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def _set_column_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _add_gap_highlighting(sheet, range_ref: str) -> None:
    if sheet.max_row < 2:
        return
    sheet.conditional_formatting.add(
        range_ref,
        CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=SUCCESS_FILL),
    )
    sheet.conditional_formatting.add(
        range_ref,
        CellIsRule(operator="greaterThan", formula=["0"], fill=CAUTION_FILL),
    )


def _parse_datetime(value: str | None) -> datetime | str:
    if not value:
        return ""
    try:
        return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return value
